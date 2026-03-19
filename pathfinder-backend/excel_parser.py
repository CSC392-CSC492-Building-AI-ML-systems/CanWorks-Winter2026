import hashlib # Python's built-in hashing library => use SHA256 to generate dedupe_hash
from datetime import datetime, date
from openpyxl import load_workbook # library to read .xlsx files
# when admin uploads an excel file via API, the file arrives as raw bytes
# BytesIO wraps the bytes so openpyxl can read them as if they were a file
from io import BytesIO


"""
This function handles the messiness of Excel dates.
Excel stores dates in 3 possible ways:
1) datetime object: call value.date() to remove the time part
2) date object: value is of desired type
3) integer: Day 1 = Dec 30, 1899. Convert by adding the integer
to the base date's day count (ordinal)
"""
def convert_excel_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return datetime.fromordinal(datetime(1899, 12, 30).toordinal() + int(value)).date()
        except Exception:
            return None
    return None

"""
This function generates the hash value to check for duplicates in the database
"""
def generate_dedupe_hash(employer, title, city):
    employer = (employer or "").lower().strip()
    title = (title or "").lower().strip()
    city = (city or "").lower().strip()
    raw = f"{employer}|{title}|{city}" # the '|' acts as the separator so that each combination of employer, title, city is unique
    return hashlib.sha256(raw.encode()).hexdigest()

def parse_excel_file(file_bytes):
    workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True)
    sheet = workbook["Main"]

    jobs = []
    errors = []

    # Row 1 is headers; data starts at row 2
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, max_col=24, values_only=True), start=2):
        try:
            # stop parsing when hit a completely empty row
            if all(cell is None for cell in row):
                break

            # Columns: A(0)=ID, B(1)=Job Title, C(2)=Employer, D(3)=Link to Posting,
            # E(4)=Application Deadline, F(5)=Mode, G(6)=Employment Type, H(7)=With Pay,
            # I(8)=Start Month, J(9)=Duration (Months), K(10)=Province, L(11)=City,
            # M(12)=Target Audience, N(13)=Description, O(14)=Responsibilities,
            # P(15)=Qualifications, Q(16)=Other Academic Requirements, R(17)=Industry,
            # S(18)=Job Function, T(19)=Seniority Level, U(20)=Compensation Min,
            # V(21)=Compensation Max, W(22)=Compensation Currency, X(23)=Skills

            if not row[1] or not row[2]:
                errors.append(f"Row {row_num}: Missing title or employer, skipped") # title and employer are not nullable
                continue

            with_pay_value = True
            if isinstance(row[7], str):
                with_pay_value = row[7].strip().lower() == "yes"
            elif isinstance(row[7], bool):
                with_pay_value = row[7]

            duration = None
            if row[9] is not None:
                try:
                    duration = float(row[9])
                except (ValueError, TypeError):
                    duration = None

            compensation_min = None
            if row[20] is not None:
                try:
                    compensation_min = float(row[20])
                except (ValueError, TypeError):
                    compensation_min = None

            compensation_max = None
            if row[21] is not None:
                try:
                    compensation_max = float(row[21])
                except (ValueError, TypeError):
                    compensation_max = None

            # Parse skills from comma-separated column X
            skills_raw = []
            if row[23]:
                skills_raw = [s.strip() for s in str(row[23]).split(",") if s.strip()]

            # create the JSON of job data
            job = {
                "title": str(row[1]).strip(),
                "employer": str(row[2]).strip(),
                "link_to_posting": row[3],
                "application_deadline": convert_excel_date(row[4]),
                "mode": row[5],
                "employment_type": row[6],
                "with_pay": with_pay_value,
                "start_month": row[8],
                "duration_months": duration,
                "province": row[10],
                "city": row[11],
                "target_audience": row[12],
                "description": row[13],
                "responsibilities": row[14],
                "qualifications": row[15],
                "other_academic_requirements": row[16],
                "industry": row[17],
                "job_function": row[18],
                "seniority_level": row[19],
                "compensation_min": compensation_min,
                "compensation_max": compensation_max,
                "compensation_currency": row[22] if row[22] else "CAD",
                "uploaded_by": "admin",
                "skills_raw": skills_raw,
                "dedupe_hash": generate_dedupe_hash(str(row[2]), str(row[1]), row[11])
            }

            # add job to list of jobs
            jobs.append(job)
        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")
    # finish parsing
    workbook.close()
    return jobs, errors
